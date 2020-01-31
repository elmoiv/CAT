# This is a dynamic Grade calculator
# Can calculate any count of subjects in the following form
# Name - Subject1 - Subject2 - ... - SubjectN - TOTAL - PERCENT
import openpyxl as xl

# Load workbook and default sheet "Sheet1"
wb = xl.load_workbook('database.xlsx')
sheet = wb['Sheet1']

# CONSTANTS
## Last column non-zero based index
MAX_C = sheet.max_column
## Global total degree
FULL_DEGREE = (MAX_C - 3) * 100

for row in range(2, sheet.max_row + 1):
    # Calculate Total degree
    total = sum(sheet.cell(row, _cell).value for _cell in range(2, sheet.max_column - 1))
    
    # Store total and percent
    sheet.cell(row, MAX_C - 1).value = total
    sheet.cell(row, MAX_C).value = total * 100 / FULL_DEGREE

# Save as new workbook
wb.save('database_processed.xlsx')
